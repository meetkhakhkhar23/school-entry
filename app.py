from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)

EXCEL_FILE = 'school_data.xlsx'

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"
    ws.append(["Name", "Class", "Roll No", "Email", "Phone"])
    wb.save(EXCEL_FILE)

@app.route('/', methods=['GET', 'POST'])
def school_form():
    if request.method == 'POST':
        name = request.form['name']
        student_class = request.form['class']
        roll = request.form['roll']
        email = request.form['email']
        phone = request.form['phone']

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, student_class, roll, email, phone])
        wb.save(EXCEL_FILE)

        return redirect('/')

    return render_template('form.html')

if __name__ == '__main__':
    app.run(debug=True)
